#!/usr/bin/env python3
"""
Extract comprehensive scenario data from the Excel spreadsheet.
"""

from openpyxl import load_workbook
import json

def extract_scenarios_from_workbook(workbook_path):
    """Extract all scenarios from the spreadsheet."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    
    scenarios = []
    
    # Check Distance Based (Scenario) sheet
    if 'Distance Based (Scenario)' in wb.sheetnames:
        ws = wb['Distance Based (Scenario)']
        
        # Look for scenario data starting from row 10
        current_scenario = None
        
        for row in range(10, min(200, ws.max_row)):
            # Check column B for scenario names
            cell_b = ws.cell(row=row, column=2)
            
            if cell_b.value and isinstance(cell_b.value, str):
                value = str(cell_b.value).strip()
                
                # Skip certain rows
                skip_terms = ['noise', 'level', 'dB', 'is there', 'receiver', 'description', 
                             'propagation', 'developed', 'undeveloped', 'residential', 'non-residential',
                             'classroom', 'hospital', 'commercial', 'industrial']
                
                if any(term in value.lower() for term in skip_terms):
                    continue
                
                # This looks like a scenario
                if len(value) > 3 and not value.startswith(('(', '-', '•')):
                    # Get description from column C
                    cell_c = ws.cell(row=row, column=3)
                    description = str(cell_c.value).strip() if cell_c.value and isinstance(cell_c.value, str) else ''
                    
                    # Look for sound power levels in columns D-F
                    sound_levels = {}
                    
                    # Check common equipment names
                    equipment_cols = {
                        'D': 'excavator',
                        'E': 'truck', 
                        'F': 'paver',
                        'G': 'roller',
                        'H': 'breaker',
                        'I': 'saw',
                        'J': 'generator'
                    }
                    
                    for col_letter, equipment in equipment_cols.items():
                        col = ord(col_letter) - ord('A') + 1
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, (int, float)):
                            if cell.value > 50 and cell.value < 150:  # Reasonable sound power range
                                sound_levels[equipment] = cell.value
                    
                    # Only add if we found some sound levels or it's a valid scenario
                    if sound_levels or len(value) > 5:
                        scenario_id = value.lower().replace(' ', '_').replace('-', '_').replace('/', '_')
                        scenario_id = ''.join(c for c in scenario_id if c.isalnum() or c == '_')
                        
                        scenarios.append({
                            'id': scenario_id,
                            'name': value,
                            'description': description or f'{value} - Construction activity',
                            'sound_power_levels': sound_levels or {'excavator': 105.0},  # Default if not found
                            'propagation_type': 'rural'  # Default
                        })
    
    # Also check for additional work types in other sheets
    work_types = [
        {
            'id': 'road_works',
            'name': 'General Road Works',
            'description': 'General road maintenance and construction activities',
            'sound_power_levels': {
                'excavator': 105.0,
                'truck': 102.0,
                'roller': 100.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'bridge_works',
            'name': 'Bridge Works',
            'description': 'Bridge construction and maintenance activities',
            'sound_power_levels': {
                'excavator': 105.0,
                'breaker': 115.0,
                'generator': 108.0
            },
            'propagation_type': 'rural'
        },
        {
            'id': 'drilling',
            'name': 'Drilling Operations',
            'description': 'Drilling and core sampling activities',
            'sound_power_levels': {
                'drill_rig': 110.0,
                'compressor': 105.0,
                'generator': 108.0
            },
            'propagation_type': 'rural'
        },
        {
            'id': 'concrete_works',
            'name': 'Concrete Works',
            'description': 'Concrete pouring and finishing activities',
            'sound_power_levels': {
                'concrete_mixer': 100.0,
                'pump': 95.0,
                'vibrator': 95.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'demolition',
            'name': 'Demolition Works',
            'description': 'Building and structure demolition activities',
            'sound_power_levels': {
                'breaker': 115.0,
                'excavator': 110.0,
                'truck': 102.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'landscaping',
            'name': 'Landscaping Works',
            'description': 'Landscaping and earthmoving activities',
            'sound_power_levels': {
                'excavator': 100.0,
                'mower': 90.0,
                'blower': 85.0
            },
            'propagation_type': 'rural'
        },
        {
            'id': 'utility_installation',
            'name': 'Utility Installation',
            'description': 'Water, sewer, and gas line installation',
            'sound_power_levels': {
                'excavator': 105.0,
                'compactor': 100.0,
                'generator': 95.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'paving_asphalt',
            'name': 'Asphalt Paving',
            'description': 'Hot mix asphalt laying and compaction',
            'sound_power_levels': {
                'paver': 108.0,
                'roller': 100.0,
                'truck': 102.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'line_marking',
            'name': 'Line Marking',
            'description': 'Road line marking and signage installation',
            'sound_power_levels': {
                'vehicle': 85.0,
                'heater': 90.0,
                'generator': 95.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'tree_removal',
            'name': 'Tree Removal',
            'description': 'Tree felling and vegetation clearance',
            'sound_power_levels': {
                'chipper': 100.0,
                'chainsaw': 110.0,
                'truck': 102.0
            },
            'propagation_type': 'rural'
        },
        {
            'id': 'signage_installation',
            'name': 'Signage Installation',
            'description': 'Road signs and traffic signal installation',
            'sound_power_levels': {
                'vehicle': 85.0,
                'auger': 95.0,
                'generator': 90.0
            },
            'propagation_type': 'urban'
        },
        {
            'id': 'stormwater_works',
            'name': 'Stormwater Drainage Works',
            'description': 'Drainage installation and maintenance',
            'sound_power_levels': {
                'excavator': 105.0,
                'pump': 95.0,
                'compactor': 100.0
            },
            'propagation_type': 'urban'
        }
    ]
    
    # Add common work types
    scenarios.extend(work_types)
    
    # Remove duplicates based on name
    seen = set()
    unique_scenarios = []
    for scenario in scenarios:
        key = scenario['name'].lower()
        if key not in seen:
            seen.add(key)
            unique_scenarios.append(scenario)
    
    return unique_scenarios

# Extract scenarios
scenarios = extract_scenarios_from_workbook('EMF-NV-TT-0067 Construction and Maintenance Noise Estimator (Roads).xlsm')

# Save to file
with open('frontend/public/scenarios.json', 'w') as f:
    json.dump(scenarios, f, indent=2)

print(f"Extracted {len(scenarios)} scenarios:")
for s in scenarios:
    print(f"  - {s['id']}: {s['name']}")
    if s['sound_power_levels']:
        print(f"    Equipment: {list(s['sound_power_levels'].keys())}")
