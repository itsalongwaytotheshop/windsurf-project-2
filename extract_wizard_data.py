#!/usr/bin/env python3
"""
Extract comprehensive information from the Excel spreadsheet for the guided wizard frontend.
"""

import json
from pathlib import Path
import sys
import os

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    print("Error: openpyxl not installed")
    exit(1)

def extract_comprehensive_data(workbook_path):
    """Extract all descriptive data, definitions, and guidance from the workbook."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    
    data = {
        "definitions": {},
        "guidance": {
            "assessment_types": {},
            "calculation_modes": {},
            "environment_approaches": {},
            "time_periods": {},
            "propagation_types": {},
            "noise_categories": {},
            "scenarios": {},
            "mitigation_measures": {
                "standard": [],
                "additional": []
            },
            "notifications": {},
            "work_hours": {},
            "compliance": {}
        },
        "images": {},
        "tips": [],
        "examples": []
    }
    
    # Extract definitions from Doc control & definitions sheet
    if 'Doc control & definitions' in wb.sheetnames:
        ws = wb['Doc control & definitions']
        for row in range(10, 50):  # Check rows 10-50 for definitions
            term = ws.cell(row=row, column=2).value
            definition = ws.cell(row=row, column=3).value
            if term and definition and isinstance(term, str) and isinstance(definition, str):
                data["definitions"][term.strip()] = definition.strip()
    
    # Extract guidance from Representative Noise Environ. sheet
    if 'Representative Noise Environ.' in wb.sheetnames:
        ws = wb['Representative Noise Environ.']
        # Extract noise category descriptions
        for row in range(10, 50):
            category_id = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            if category_id and description:
                data["guidance"]["noise_categories"][str(category_id)] = {
                    "description": str(description),
                    "examples": []
                }
    
    # Extract scenario information
    if 'Distance Based (Scenario)' in wb.sheetnames:
        ws = wb['Distance Based (Scenario)']
        # Look for scenario descriptions
        for row in range(10, 100):
            scenario_name = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            if scenario_name and description:
                data["guidance"]["scenarios"][str(scenario_name)] = {
                    "description": str(description),
                    "typical_activities": []
                }
    
    # Extract mitigation measures
    if 'Standard Measures' in wb.sheetnames:
        ws = wb['Standard Measures']
        for row in range(10, 100):
            measure = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            if measure and description:
                data["guidance"]["mitigation_measures"]["standard"].append({
                    "title": str(measure),
                    "description": str(description),
                    "reduction": ws.cell(row=row, column=4).value or 0
                })
    
    if 'Additional Measures' in wb.sheetnames:
        ws = wb['Additional Measures']
        for row in range(10, 100):
            measure = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            if measure and description:
                data["guidance"]["mitigation_measures"]["additional"].append({
                    "title": str(measure),
                    "description": str(description),
                    "reduction": ws.cell(row=row, column=4).value or 0
                })
    
    # Extract notification requirements
    if 'Factsheet (maintenance)' in wb.sheetnames:
        ws = wb['Factsheet (maintenance)']
        # Extract notification methods and requirements
        for row in range(10, 100):
            method = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=4).value
            if method and description:
                data["guidance"]["notifications"][str(method)] = {
                    "description": str(description),
                    "timing": ws.cell(row=row, column=3).value or ""
                }
    
    # Extract work hour restrictions
    if 'Distance Based Summary' in wb.sheetnames:
        ws = ws = wb['Distance Based Summary']
        # Look for work hour guidance
        for row in range(10, 50):
            period = ws.cell(row=row, column=2).value
            restriction = ws.cell(row=row, column=3).value
            if period and restriction:
                data["guidance"]["work_hours"][str(period)] = str(restriction)
    
    # Add comprehensive guidance text
    data["guidance"]["assessment_types"] = {
        "distance_based": {
            "title": "Distance Based Assessment",
            "description": "Calculate noise levels at a specific distance from the source. This is the most common assessment type for determining if noise levels at nearby properties will exceed criteria.",
            "when_to_use": "Use when you need to know the noise level at a specific location (e.g., a residential property) at a known distance from the work site.",
            "steps": [
                "Select the work scenario or plant",
                "Enter the distance to the receiver",
                "Choose the time period",
                "Review results and mitigation measures"
            ]
        },
        "full_estimator": {
            "title": "Full Estimator Assessment",
            "description": "Comprehensive assessment that calculates affected distances and provides detailed analysis including respite periods and notification requirements.",
            "when_to_use": "Use for detailed project planning, when you need to determine the full extent of noise impact and all compliance requirements.",
            "steps": [
                "Select the work scenario or plant",
                "Specify the work location",
                "Choose time periods for work",
                "Review comprehensive results including notifications and compliance"
            ]
        }
    }
    
    data["guidance"]["calculation_modes"] = {
        "scenario": {
            "title": "Scenario-based Calculation",
            "description": "Uses predefined work scenarios (e.g., excavation, paving) with typical equipment combinations.",
            "when_to_use": "Use when your work matches one of the standard scenarios. This provides the most accurate results for common construction activities."
        },
        "noisiest_plant": {
            "title": "Noisiest Plant Calculation",
            "description": "Identifies and uses the noisiest piece of equipment from your selected scenario.",
            "when_to_use": "Use when you want a conservative estimate based on the single noisiest piece of equipment."
        },
        "individual_plant": {
            "title": "Individual Plant Selection",
            "description": "Allows you to select specific pieces of equipment for custom scenarios.",
            "when_to_use": "Use when your work involves a specific combination of equipment that doesn't match standard scenarios."
        }
    }
    
    data["guidance"]["environment_approaches"] = {
        "representative_noise_environment": {
            "title": "Representative Noise Environment",
            "description": "Uses predefined background noise levels based on the noise category (R1, U2, etc.).",
            "when_to_use": "Use for most assessments where you don't have measured background noise levels."
        },
        "user_supplied_background_level": {
            "title": "User Supplied Background Level",
            "description": "Allows you to enter a measured or known background noise level.",
            "when_to_use": "Use when you have actual measurements of background noise at the site."
        }
    }
    
    data["guidance"]["time_periods"] = {
        "day": {
            "title": "Daytime (7am - 6pm)",
            "description": "Standard daytime work hours. Most construction activities occur during this period.",
            "restrictions": "Monday to Saturday only. No work on Sundays or public holidays (except emergency works).",
            "typical_background": "45 dB for R1 (Rural Residential), 55 dB for U2 (Urban Industrial)"
        },
        "evening": {
            "title": "Evening (6pm - 10pm)",
            "description": "Evening work hours with additional restrictions.",
            "restrictions": "Maximum 3 consecutive days, EPA approval required for noisy work.",
            "typical_background": "40 dB for R1 (Rural Residential), 50 dB for U2 (Urban Industrial)"
        },
        "night": {
            "title": "Nighttime (10pm - 7am)",
            "description": "Night work hours with the strictest requirements.",
            "restrictions": "Emergency works only, maximum 2 consecutive nights, requires EPA Section 115 approval.",
            "typical_background": "35 dB for R1 (Rural Residential), 45 dB for U2 (Urban Industrial)"
        }
    }
    
    data["guidance"]["propagation_types"] = {
        "rural": {
            "title": "Rural",
            "description": "Open country environment with minimal obstacles. Noise propagates with minimal attenuation.",
            "characteristics": "Flat or gently rolling terrain, scattered vegetation, no buildings",
            "example": "Open farmland, rural road works"
        },
        "urban": {
            "title": "Urban",
            "description": "Built-up environment with buildings and obstacles that affect noise propagation.",
            "characteristics": "Buildings, streets, urban infrastructure, reflective surfaces",
            "example": "City streets, suburban areas"
        },
        "hard_ground": {
            "title": "Hard Ground",
            "description": "Solid, reflective ground surface that provides less noise attenuation.",
            "characteristics": "Concrete, asphalt, compacted earth, rock",
            "example": "Paved areas, concrete pads, hard-standing"
        },
        "soft_ground": {
            "title": "Soft Ground",
            "description": "Porous ground surface that provides additional noise attenuation.",
            "characteristics": "Grass, soil, sand, vegetation",
            "example": "Parks, gardens, unpaved areas"
        },
        "mixed": {
            "title": "Mixed",
            "description": "Combination of different ground types and environments.",
            "characteristics": "Variable terrain with both hard and soft surfaces",
            "example": "Suburban areas with gardens and driveways"
        }
    }
    
    # Add tips and best practices
    data["tips"] = [
        "Always measure background noise at the most sensitive receiver location",
        "Consider the worst-case scenario when planning work hours",
        "Implement mitigation measures before considering out-of-hours work",
        "Keep records of all notifications and complaints",
        "Use noise monitoring for high-impact projects",
        "Consider the cumulative impact of multiple noise sources",
        "Plan respite periods for extended out-of-hours work"
    ]
    
    wb.close()
    return data

# Extract and save the data
workbook_path = "EMF-NV-TT-0067 Construction and Maintenance Noise Estimator (Roads).xlsm"
if Path(workbook_path).exists():
    data = extract_comprehensive_data(workbook_path)
    
    # Save as JSON for the frontend
    with open('frontend/public/wizard-data.json', 'w') as f:
        json.dump(data, f, indent=2)
    
    print("Successfully extracted comprehensive data from spreadsheet")
    print(f"Definitions: {len(data['definitions'])}")
    print(f"Scenarios: {len(data['guidance']['scenarios'])}")
    print(f"Standard Measures: {len(data['guidance']['mitigation_measures']['standard'])}")
    print(f"Additional Measures: {len(data['guidance']['mitigation_measures']['additional'])}")
else:
    print(f"Workbook not found: {workbook_path}")
