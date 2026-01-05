#!/usr/bin/env python3
"""
Workbook mapper script to analyze the Excel macro workbook structure.
This script maps sheets, tables, defined names, and identifies key data structures.
"""

import json
from pathlib import Path
from typing import Dict, List, Any, Optional
import hashlib

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)


def analyze_workbook(workbook_path: str) -> Dict[str, Any]:
    """Analyze the workbook structure and return comprehensive mapping."""
    print(f"Analyzing workbook: {workbook_path}")
    
    # Calculate file hash
    with open(workbook_path, 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()
    
    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    
    analysis = {
        "workbook_info": {
            "file_path": str(workbook_path),
            "file_name": Path(workbook_path).name,
            "file_hash": file_hash,
            "sheet_count": len(wb.sheetnames),
        },
        "sheets": {},
        "defined_names": {},
        "candidate_tables": {}
    }
    
    # Analyze defined names
    for name, defined_name in wb.defined_names.items():
        try:
            analysis["defined_names"][name] = {
                "attr_text": str(defined_name.attr_text) if hasattr(defined_name, 'attr_text') else str(defined_name),
                "value": str(defined_name.value) if hasattr(defined_name, 'value') else None,
                "type": type(defined_name).__name__
            }
        except Exception as e:
            analysis["defined_names"][name] = {
                "error": str(e),
                "type": type(defined_name).__name__
            }
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "title": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimensions": ws.calculate_dimension() if hasattr(ws, 'calculate_dimension') else f"A1:{ws.max_row}:{ws.max_column}",
            "hidden": ws.sheet_state == 'hidden',
            "tables_detected": [],
            "data_ranges": [],
            "formulas": [],
            "constants": []
        }
        
        # Look for potential tables (contiguous data blocks with headers)
        if ws.max_row > 0 and ws.max_column > 0:
            tables = detect_tables(ws)
            sheet_info["tables_detected"] = tables
        
        # Look for data ranges (non-empty cells)
        data_ranges = find_data_ranges(ws)
        sheet_info["data_ranges"] = data_ranges
        
        # Sample some cell content to understand the sheet
        sample_content = sample_sheet_content(ws)
        sheet_info["sample_content"] = sample_content
        
        analysis["sheets"][sheet_name] = sheet_info
    
    # Identify candidate sheets for different purposes
    analysis["candidate_sheets"] = identify_candidate_sheets(analysis["sheets"])
    
    wb.close()
    return analysis


def detect_tables(ws) -> List[Dict[str, Any]]:
    """Detect potential tables in a worksheet."""
    tables = []
    
    # Look for ranges that might be tables
    # Start by finding header rows (cells with text content)
    for row in range(1, min(ws.max_row + 1, 50)):  # Check first 50 rows for headers
        header_cells = []
        for col in range(1, min(ws.max_column + 1, 20)):  # Check first 20 columns
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).strip():
                header_cells.append((col, cell.value))
        
        # If we found a potential header row, see if there's data below
        if len(header_cells) >= 2:  # At least 2 columns
            start_col = header_cells[0][0]
            end_col = header_cells[-1][0]
            
            # Check if there's data in the next few rows
            data_rows = 0
            for check_row in range(row + 1, min(row + 11, ws.max_row + 1)):
                has_data = False
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=check_row, column=col)
                    if cell.value is not None:
                        has_data = True
                        break
                if has_data:
                    data_rows += 1
            
            if data_rows > 0:
                start_cell = ws.cell(row=row, column=start_col)
                end_cell = ws.cell(row=row+data_rows, column=end_col)
                
                # Handle empty cells
                start_coord = start_cell.coordinate if hasattr(start_cell, 'coordinate') else f"A{row}"
                end_coord = end_cell.coordinate if hasattr(end_cell, 'coordinate') else f"{chr(64+end_col)}{row+data_rows}"
                
                table_range = f"{start_coord}:{end_coord}"
                tables.append({
                    "header_row": row,
                    "start_col": start_col,
                    "end_col": end_col,
                    "data_rows": data_rows,
                    "range": table_range,
                    "headers": [cell[1] for cell in header_cells]
                })
    
    return tables


def find_data_ranges(ws) -> List[Dict[str, Any]]:
    """Find contiguous data ranges in the worksheet."""
    ranges = []
    
    # Simple approach: find non-empty cell clusters
    used_range = ws.calculate_dimension()
    if used_range:
        ranges.append({
            "type": "used_range",
            "range": used_range,
            "description": "Worksheet used range"
        })
    
    return ranges


def sample_sheet_content(ws) -> Dict[str, Any]:
    """Sample content from different parts of the sheet."""
    samples = {
        "first_row": [],
        "first_column": [],
        "center_area": []
    }
    
    # Sample first row
    for col in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=1, column=col)
        if cell.value is not None:
            samples["first_row"].append(str(cell.value))
    
    # Sample first column
    for row in range(1, min(ws.max_row + 1, 10)):
        cell = ws.cell(row=row, column=1)
        if cell.value is not None:
            samples["first_column"].append(str(cell.value))
    
    # Sample center area (rows 5-15, cols 3-8)
    for row in range(5, min(16, ws.max_row + 1)):
        row_data = []
        for col in range(3, min(9, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                row_data.append(str(cell.value))
            else:
                row_data.append("")
        if any(row_data):
            samples["center_area"].append(row_data)
    
    return samples


def identify_candidate_sheets(sheets: Dict[str, Any]) -> Dict[str, List[str]]:
    """Identify sheets that likely contain specific types of data."""
    candidates = {
        "scenarios": [],
        "plants": [],
        "categories": [],
        "measures": [],
        "distance_tables": [],
        "worked_examples": [],
        "background_data": [],
        "constants": []
    }
    
    for sheet_name, sheet_info in sheets.items():
        sample = sheet_info.get("sample_content", {})
        tables = sheet_info.get("tables_detected", [])
        
        # Look for scenario-related content
        if any("scenario" in str(cell).lower() for row in sample.get("first_row", []) for cell in [row]):
            candidates["scenarios"].append(sheet_name)
        elif any("scenario" in str(cell).lower() for row in sample.get("first_column", []) for cell in [row]):
            candidates["scenarios"].append(sheet_name)
        
        # Look for plant/source content
        if any("plant" in str(cell).lower() or "source" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["plants"].append(sheet_name)
        
        # Look for category/noise area content
        if any("categor" in str(cell).lower() or "noise" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["categories"].append(sheet_name)
        
        # Look for measures/mitigation content
        if any("measure" in str(cell).lower() or "mitigat" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["measures"].append(sheet_name)
        
        # Look for distance-related content
        if any("distance" in str(cell).lower() or "concawe" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["distance_tables"].append(sheet_name)
        
        # Look for worked examples
        if any("example" in str(cell).lower() or "worked" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["worked_examples"].append(sheet_name)
        
        # Look for background/environment data
        if any("background" in str(cell).lower() or "environment" in str(cell).lower() 
               for row in sample.get("first_row", []) for cell in [row]):
            candidates["background_data"].append(sheet_name)
    
    return candidates


def save_analysis(analysis: Dict[str, Any], output_path: str):
    """Save analysis to JSON and markdown files."""
    # Save JSON
    json_path = Path(output_path).with_suffix('.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, default=str)
    
    # Save markdown summary
    md_path = Path(output_path).with_suffix('.md')
    with open(md_path, 'w', encoding='utf-8') as f:
        write_markdown_summary(analysis, f)
    
    print(f"Analysis saved to: {json_path}")
    print(f"Summary saved to: {md_path}")


def write_markdown_summary(analysis: Dict[str, Any], f):
    """Write a markdown summary of the workbook analysis."""
    f.write("# Workbook Analysis Summary\n\n")
    
    # Workbook info
    info = analysis["workbook_info"]
    f.write(f"## Workbook Information\n")
    f.write(f"- **File**: {info['file_name']}\n")
    f.write(f"- **Hash**: {info['file_hash'][:16]}...\n")
    f.write(f"- **Sheets**: {info['sheet_count']}\n\n")
    
    # Sheets overview
    f.write("## Sheets Overview\n\n")
    for sheet_name, sheet_info in analysis["sheets"].items():
        status = "🔒" if sheet_info["hidden"] else "📄"
        f.write(f"{status} **{sheet_name}** ({sheet_info['max_row']}x{sheet_info['max_column']})\n")
        
        if sheet_info["tables_detected"]:
            f.write(f"  - Tables: {len(sheet_info['tables_detected'])}\n")
        
        # Show sample content hints
        sample = sheet_info.get("sample_content", {})
        if sample.get("first_row"):
            f.write(f"  - Headers: {sample['first_row'][:3]}...\n")
    
    # Candidate sheets
    f.write("\n## Candidate Sheets by Purpose\n\n")
    candidates = analysis["candidate_sheets"]
    for purpose, sheet_list in candidates.items():
        if sheet_list:
            f.write(f"### {purpose.replace('_', ' ').title()}\n")
            for sheet in sheet_list:
                f.write(f"- {sheet}\n")
            f.write("\n")
    
    # Defined names
    if analysis["defined_names"]:
        f.write("## Defined Names\n\n")
        for name, info in analysis["defined_names"].items():
            f.write(f"- **{name}**: {info.get('attr_text', 'N/A')}\n")
    
    # Tables detected
    f.write("\n## Tables Detected\n\n")
    total_tables = sum(len(sheet["tables_detected"]) for sheet in analysis["sheets"].values())
    f.write(f"Total tables found: {total_tables}\n\n")
    
    for sheet_name, sheet_info in analysis["sheets"].items():
        tables = sheet_info["tables_detected"]
        if tables:
            f.write(f"### {sheet_name}\n")
            for i, table in enumerate(tables, 1):
                f.write(f"**Table {i}** ({table['range']}):\n")
                f.write(f"- Headers: {table['headers'][:5]}...\n")
                f.write(f"- Data rows: {table['data_rows']}\n\n")


def main():
    """Main function to run the workbook mapper."""
    workbook_path = "EMF-NV-TT-0067 Construction and Maintenance Noise Estimator (Roads).xlsm"
    
    if not Path(workbook_path).exists():
        print(f"Error: Workbook not found at {workbook_path}")
        print("Please ensure the workbook file is accessible.")
        return
    
    try:
        analysis = analyze_workbook(workbook_path)
        save_analysis(analysis, "docs/workbook_map")
        
        print("\nWorkbook mapping completed successfully!")
        print("Key findings:")
        
        # Print key statistics
        info = analysis["workbook_info"]
        print(f"- {info['sheet_count']} sheets found")
        print(f"- {len(analysis['defined_names'])} defined names")
        
        total_tables = sum(len(sheet["tables_detected"]) for sheet in analysis["sheets"].values())
        print(f"- {total_tables} potential tables detected")
        
        candidates = analysis["candidate_sheets"]
        print(f"- Potential scenario sheets: {len(candidates['scenarios'])}")
        print(f"- Potential plant sheets: {len(candidates['plants'])}")
        print(f"- Potential category sheets: {len(candidates['categories'])}")
        print(f"- Potential worked examples: {len(candidates['worked_examples'])}")
        
    except Exception as e:
        print(f"Error analyzing workbook: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
