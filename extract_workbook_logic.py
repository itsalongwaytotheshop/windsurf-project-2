#!/usr/bin/env python3
"""
Advanced workbook extraction to capture calculation logic and formulas.
"""

import json
import hashlib
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    raise ImportError("openpyxl is required for dataset extraction")

logger = logging.getLogger(__name__)


def extract_workbook_logic(workbook_path: str, output_path: str):
    """Extract detailed logic and formulas from the workbook."""
    workbook_path = Path(workbook_path)
    output_path = Path(output_path)
    
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    
    # Load workbook with formulas
    wb = load_workbook(
        workbook_path, 
        read_only=False,  # Need formulas
        data_only=False,  # Get formulas, not just values
        keep_links=False
    )
    
    extraction = {
        "workbook_info": {
            "file_name": workbook_path.name,
            "file_hash": hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        },
        "sheets": {},
        "key_logic": {}
    }
    
    # Extract key sheets with formulas
    key_sheets = [
        "Estimator (Scenario)",
        "Estimator (Individual Plant)",
        "Distance Based (Scenario)",
        "Distance Based (Noisiest Plant)",
        "Scenario_SWL",
        "conc_scen",
        "Concawe",
        "Representative Noise Environ.",
        "Worked Examples"
    ]
    
    for sheet_name in key_sheets:
        if sheet_name in wb.sheetnames:
            extraction["sheets"][sheet_name] = extract_sheet_with_formulas(wb[sheet_name])
    
    # Extract specific calculation logic
    extraction["key_logic"] = {
        "concawe_tables": extract_concawe_tables(wb),
        "scenario_swl": extract_scenario_swl(wb),
        "noise_categories": extract_noise_categories(wb),
        "worked_examples": extract_worked_examples_detailed(wb),
        "propagation_formulas": extract_propagation_formulas(wb),
        "distance_calculations": extract_distance_calculations(wb)
    }
    
    # Save extraction
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(extraction, f, indent=2, default=str)
    
    print(f"Logic extraction saved to: {output_path}")
    wb.close()


def extract_sheet_with_formulas(ws):
    """Extract sheet data including formulas."""
    sheet_data = {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "formulas": {},
        "named_ranges": {},
        "key_cells": {}
    }
    
    # Extract formulas from key areas
    formula_ranges = [
        "A1:Z100",  # General area
        "A1:AZ1000" if ws.max_column > 26 else "A1:Z100"
    ]
    
    for range_str in formula_ranges:
        for row in ws[range_str]:
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    cell_ref = f"{cell.column_letter}{cell.row}"
                    sheet_data["formulas"][cell_ref] = {
                        "formula": cell.value,
                        "value": cell.displayed_value if hasattr(cell, 'displayed_value') else None
                    }
    
    # Look for key calculation cells
    key_patterns = ["LAeq", "SWL", "NML", "RBL", "distance", "level"]
    for row in ws.iter_rows(max_row=100, max_col=50):
        for cell in row:
            if cell.value and any(pattern.lower() in str(cell.value).lower() for pattern in key_patterns):
                cell_ref = f"{cell.column_letter}{cell.row}"
                sheet_data["key_cells"][cell_ref] = {
                    "value": cell.value,
                    "type": cell.data_type,
                    "formula": ws.cell(row=cell.row, column=cell.column + 1).value if cell.column < ws.max_column else None
                }
    
    return sheet_data


def extract_concawe_tables(wb):
    """Extract Concawe propagation tables."""
    concawe_data = {}
    
    # Check for conc_scen sheet
    if "conc_scen" in wb.sheetnames:
        ws = wb["conc_scen"]
        
        # Look for distance vs level data
        concawe_data["distance_levels"] = []
        
        # Extract the table (likely starts around row 2-3)
        for row in range(2, min(100, ws.max_row + 1)):
            row_data = []
            has_data = False
            
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    has_data = True
                    row_data.append(cell.value)
                else:
                    row_data.append(None)
            
            if has_data and len([x for x in row_data if x is not None]) > 2:
                concawe_data["distance_levels"].append(row_data)
    
    # Check for Concawe sheet
    if "Concawe" in wb.sheetnames:
        ws = wb["Concawe"]
        concawe_data["attenuation_table"] = []
        
        # Extract attenuation data
        for row in range(2, min(100, ws.max_row + 1)):
            row_data = []
            has_data = False
            
            for col in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    has_data = True
                    row_data.append(cell.value)
                else:
                    row_data.append(None)
            
            if has_data:
                concawe_data["attenuation_table"].append(row_data)
    
    return concawe_data


def extract_scenario_swl(wb):
    """Extract scenario sound power levels."""
    scenario_data = {}
    
    if "Scenario_SWL" in wb.sheetnames:
        ws = wb["Scenario_SWL"]
        
        # Look for scenario names and SWL values
        scenarios = {}
        
        # Find header row
        header_row = None
        for row in range(1, min(10, ws.max_row + 1)):
            first_cell = ws.cell(row=row, column=1)
            if first_cell.value and "rural" in str(first_cell.value).lower():
                header_row = row
                break
        
        if header_row:
            # Get column headers (scenarios)
            headers = []
            for col in range(2, min(ws.max_column + 1, 60)):
                cell = ws.cell(row=header_row, column=col)
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            # Extract SWL data for each scenario
            for col_idx, header in enumerate(headers, start=2):
                scenario_swls = {}
                
                for row in range(header_row + 1, min(header_row + 60, ws.max_row + 1)):
                    plant_cell = ws.cell(row=row, column=1)
                    swl_cell = ws.cell(row=row, column=col_idx)
                    
                    if plant_cell.value and swl_cell.value:
                        plant_name = str(plant_cell.value)
                        swl = float(swl_cell.value) if isinstance(swl_cell.value, (int, float)) else None
                        if swl is not None:
                            scenario_swls[plant_name] = swl
                
                if scenario_swls:
                    scenarios[header] = scenario_swls
        
        scenario_data["scenarios"] = scenarios
    
    return scenario_data


def extract_noise_categories(wb):
    """Extract noise area categories and background levels."""
    category_data = {}
    
    if "Representative Noise Environ." in wb.sheetnames:
        ws = wb["Representative Noise Environ."]
        
        # Look for the category table
        categories = {}
        
        # Find the RBL table
        for row in range(1, min(50, ws.max_row + 1)):
            first_cell = ws.cell(row=row, column=1)
            if first_cell.value and "RBL" in str(first_cell.value):
                # Found the header, extract data
                headers = []
                for col in range(2, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        headers.append(str(cell.value))
                    else:
                        break
                
                # Extract time period data
                time_periods = ["Day", "Evening", "Night"]
                for period_idx, period in enumerate(time_periods):
                    if row + period_idx + 1 <= ws.max_row:
                        period_data = {}
                        for col_idx, header in enumerate(headers, start=2):
                            cell = ws.cell(row=row + period_idx + 1, column=col_idx)
                            if cell.value is not None:
                                period_data[header] = float(cell.value)
                        
                        categories[period] = period_data
                
                break
        
        category_data["background_levels"] = categories
    
    return category_data


def extract_worked_examples_detailed(wb):
    """Extract worked examples with inputs and expected outputs."""
    examples = []
    
    if "Worked Examples" in wb.sheetnames:
        ws = wb["Worked Examples"]
        
        # Look for example blocks
        current_example = None
        
        for row in range(1, min(300, ws.max_row + 1)):
            first_cell = ws.cell(row=row, column=1)
            
            # Look for example markers
            if first_cell.value and "example" in str(first_cell.value).lower():
                if current_example:
                    examples.append(current_example)
                
                current_example = {
                    "title": str(first_cell.value),
                    "inputs": {},
                    "outputs": {},
                    "description": None
                }
            
            # Extract input/output pairs
            elif current_example:
                for col in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    label_cell = ws.cell(row=row - 1, column=col) if row > 1 else None
                    
                    if cell.value and label_cell and label_cell.value:
                        label = str(label_cell.value).strip()
                        value = cell.value
                        
                        # Classify as input or output
                        if any(keyword in label.lower() for keyword in ["input", "given", "selected", "category", "distance"]):
                            current_example["inputs"][label] = value
                        elif any(keyword in label.lower() for keyword in ["output", "result", "level", "exceed", "distance"]):
                            current_example["outputs"][label] = value
        
        if current_example:
            examples.append(current_example)
    
    return {"worked_examples": examples}


def extract_propagation_formulas(wb):
    """Extract propagation calculation formulas."""
    formulas = {}
    
    # Check Estimator sheets for propagation formulas
    for sheet_name in ["Estimator (Scenario)", "Estimator (Individual Plant)"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Look for cells with propagation calculations
            for row in range(1, min(200, ws.max_row + 1)):
                for col in range(1, min(50, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    
                    if cell.data_type == 'f':  # Formula cell
                        formula_str = str(cell.value)
                        
                        # Look for propagation-related formulas
                        if any(keyword in formula_str.lower() for keyword in ["log", "distance", "attenu", "concawe"]):
                            cell_ref = f"{cell.column_letter}{cell.row}"
                            formulas[f"{sheet_name}_{cell_ref}"] = {
                                "formula": formula_str,
                                "context": f"Row {row}, Col {col}"
                            }
    
    return formulas


def extract_distance_calculations(wb):
    """Extract distance calculation methodology."""
    distance_methods = {}
    
    # Check distance-based sheets
    for sheet_name in ["Distance Based (Scenario)", "Distance Based (Noisiest Plant)"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            methods = {}
            
            # Look for distance calculation logic
            for row in range(1, min(100, ws.max_row + 1)):
                for col in range(1, min(30, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    
                    if cell.data_type == 'f':  # Formula cell
                        formula_str = str(cell.value)
                        
                        # Look for goal seek or inverse calculations
                        if any(keyword in formula_str.lower() for keyword in ["goal", "seek", "inverse", "solve"]):
                            cell_ref = f"{cell.column_letter}{cell.row}"
                            methods[f"{sheet_name}_{cell_ref}"] = {
                                "formula": formula_str,
                                "type": "inverse_calculation"
                            }
            
            distance_methods[sheet_name] = methods
    
    return distance_methods


def main():
    """Main extraction function."""
    workbook_path = "EMF-NV-TT-0067 Construction and Maintenance Noise Estimator (Roads).xlsm"
    output_path = "docs/workbook_logic.json"
    
    try:
        extract_workbook_logic(workbook_path, output_path)
        print("Workbook logic extraction completed successfully!")
        
    except Exception as e:
        print(f"Error extracting workbook logic: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
