"""
Dataset extraction module for reading Excel workbooks and converting to JSON datasets.
"""

import json
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    raise ImportError("openpyxl is required for dataset extraction")

logger = logging.getLogger(__name__)


class DatasetExtractor:
    """Extracts data from Excel workbooks and converts to JSON datasets."""
    
    def __init__(self):
        self.workbook = None
        self.defined_names = {}
    
    def extract_dataset(self, workbook_path: Union[str, Path], output_dir: Union[str, Path]) -> str:
        """Extract dataset from Excel workbook.
        
        Args:
            workbook_path: Path to the Excel workbook file.
            output_dir: Directory to save the extracted dataset.
            
        Returns:
            Version string of the extracted dataset.
        """
        workbook_path = Path(workbook_path)
        output_dir = Path(output_dir)
        
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        # Calculate workbook hash
        with open(workbook_path, 'rb') as f:
            workbook_hash = hashlib.sha256(f.read()).hexdigest()
        
        # Load workbook
        logger.info(f"Loading workbook: {workbook_path}")
        self.workbook = load_workbook(
            workbook_path, 
            read_only=True, 
            data_only=True, 
            keep_links=False
        )
        
        # Extract defined names
        self._extract_defined_names()
        
        # Extract all relevant tables
        tables = {}
        table_metadata = {}
        
        # Define extraction mappings
        extraction_plan = self._get_extraction_plan()
        
        for table_name, extraction_config in extraction_plan.items():
            try:
                table_data, metadata = self._extract_table(extraction_config)
                if table_data:
                    tables[table_name] = table_data
                    table_metadata[table_name] = metadata
                    logger.info(f"Extracted {len(table_data)} rows from {table_name}")
                else:
                    logger.warning(f"No data extracted for {table_name}")
                    
            except Exception as e:
                logger.error(f"Failed to extract {table_name}: {e}")
                continue
        
        # Create dataset metadata
        version = self._generate_version(workbook_hash)
        dataset_metadata = {
            "workbook_name": workbook_path.name,
            "extraction_timestamp": datetime.utcnow().isoformat(),
            "workbook_hash": workbook_hash,
            "version": version,
            "total_tables": len(tables),
            "sheet_count": len(self.workbook.sheetnames)
        }
        
        # Create complete dataset
        dataset = {
            "metadata": dataset_metadata,
            "tables": tables,
            "table_metadata": table_metadata
        }
        
        # Save dataset
        version_dir = output_dir / version
        version_dir.mkdir(parents=True, exist_ok=True)
        
        dataset_file = version_dir / "dataset.json"
        with open(dataset_file, 'w', encoding='utf-8') as f:
            json.dump(dataset, f, indent=2, default=str)
        
        # Save extraction report
        report_file = version_dir / "extraction_report.json"
        extraction_report = self._create_extraction_report(dataset_metadata, table_metadata)
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(extraction_report, f, indent=2, default=str)
        
        self.workbook.close()
        
        logger.info(f"Dataset extracted to {dataset_file}")
        return version
    
    def _extract_defined_names(self):
        """Extract defined names from workbook."""
        self.defined_names = {}
        
        for name in self.workbook.defined_names.definedName:
            try:
                self.defined_names[name.name] = {
                    "attr_text": name.attr_text,
                    "value": name.value if hasattr(name, 'value') else None,
                    "type": type(name).__name__
                }
            except Exception as e:
                logger.warning(f"Failed to extract defined name {name.name}: {e}")
    
    def _get_extraction_plan(self) -> Dict[str, Dict[str, Any]]:
        """Define the extraction plan for workbook tables."""
        return {
            "noise_categories": {
                "sheet_patterns": ["categor", "noise", "area"],
                "required_columns": ["id", "name"],
                "data_type": "list"
            },
            "scenarios": {
                "sheet_patterns": ["scenario", "swl"],
                "required_columns": ["id", "name", "sound_power_level"],
                "data_type": "list"
            },
            "plants": {
                "sheet_patterns": ["plant", "source", "equipment"],
                "required_columns": ["id", "name", "sound_power_level"],
                "data_type": "list"
            },
            "mitigation_measures": {
                "sheet_patterns": ["measure", "mitigat"],
                "required_columns": ["id", "title", "text", "type"],
                "data_type": "list"
            },
            "background_levels": {
                "sheet_patterns": ["background", "environment", "rbl"],
                "required_columns": ["category", "time_period", "level"],
                "data_type": "list"
            },
            "propagation_data": {
                "sheet_patterns": ["propagat", "concawe", "distance"],
                "required_columns": ["distance", "attenuation"],
                "data_type": "list"
            },
            "distance_tables": {
                "sheet_patterns": ["distance", "concawe", "appendix"],
                "required_columns": ["distance", "level"],
                "data_type": "list"
            },
            "worked_examples": {
                "sheet_patterns": ["example", "worked", "test"],
                "required_columns": ["input", "expected_output"],
                "data_type": "list"
            }
        }
    
    def _extract_table(self, config: Dict[str, Any]) -> Tuple[List[Any], Dict[str, Any]]:
        """Extract a specific table based on configuration."""
        sheet_patterns = config["sheet_patterns"]
        required_columns = config["required_columns"]
        data_type = config["data_type"]
        
        # Find matching sheets
        matching_sheets = self._find_matching_sheets(sheet_patterns)
        
        for sheet_name in matching_sheets:
            try:
                ws = self.workbook[sheet_name]
                
                # Detect tables in the sheet
                tables = self._detect_tables(ws, required_columns)
                
                if tables:
                    # Use the first suitable table found
                    table = tables[0]
                    data = self._extract_table_data(ws, table, required_columns)
                    
                    metadata = {
                        "sheet_name": sheet_name,
                        "cell_range": table["range"],
                        "column_types": self._infer_column_types(data),
                        "row_count": len(data),
                        "has_headers": True
                    }
                    
                    return data, metadata
                
            except Exception as e:
                logger.warning(f"Failed to extract from sheet {sheet_name}: {e}")
                continue
        
        return [], {}
    
    def _find_matching_sheets(self, patterns: List[str]) -> List[str]:
        """Find sheets matching the given patterns."""
        matching = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet_lower = sheet_name.lower()
            
            if any(pattern.lower() in sheet_lower for pattern in patterns):
                matching.append(sheet_name)
        
        return matching
    
    def _detect_tables(self, ws, required_columns: List[str]) -> List[Dict[str, Any]]:
        """Detect tables in a worksheet."""
        tables = []
        
        # Look for header rows
        for row in range(1, min(ws.max_row + 1, 50)):
            header_cells = []
            for col in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    header_cells.append((col, str(cell.value).strip()))
            
            # Check if this looks like a header row with required columns
            header_texts = [cell[1] for cell in header_cells]
            if self._matches_required_columns(header_texts, required_columns):
                
                # Find data extent
                start_col = header_cells[0][0]
                end_col = header_cells[-1][0]
                
                # Count data rows
                data_rows = 0
                for check_row in range(row + 1, min(row + 101, ws.max_row + 1)):
                    has_data = False
                    for col in range(start_col, end_col + 1):
                        cell = ws.cell(row=check_row, column=col)
                        if cell.value is not None:
                            has_data = True
                            break
                    if has_data:
                        data_rows += 1
                    elif data_rows > 0:
                        # Stop at first empty row after finding data
                        break
                
                if data_rows > 0:
                    table_range = f"{ws.cell(row=row, column=start_col).coordinate}:{ws.cell(row=row+data_rows, column=end_col).coordinate}"
                    tables.append({
                        "header_row": row,
                        "start_col": start_col,
                        "end_col": end_col,
                        "data_rows": data_rows,
                        "range": table_range,
                        "headers": header_texts
                    })
        
        return tables
    
    def _matches_required_columns(self, headers: List[str], required: List[str]) -> bool:
        """Check if headers match required columns."""
        header_lower = [h.lower() for h in headers]
        required_lower = [r.lower() for r in required]
        
        # Check if all required columns are present (allowing partial matches)
        matches = 0
        for req in required_lower:
            if any(req in h for h in header_lower):
                matches += 1
        
        return matches >= len(required) * 0.7  # Allow 70% match
    
    def _extract_table_data(self, ws, table: Dict[str, Any], required_columns: List[str]) -> List[Dict[str, Any]]:
        """Extract data from a detected table."""
        data = []
        headers = table["headers"]
        header_row = table["header_row"]
        start_col = table["start_col"]
        end_col = table["end_col"]
        data_rows = table["data_rows"]
        
        # Create column mapping
        col_mapping = {}
        for i, header in enumerate(headers):
            col_mapping[start_col + i] = header
        
        # Extract data rows
        for row_offset in range(1, data_rows + 1):
            row_data = {}
            has_data = False
            
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=header_row + row_offset, column=col)
                value = cell.value
                
                if value is not None:
                    has_data = True
                
                # Convert to appropriate type
                if isinstance(value, str):
                    value = value.strip()
                    if value == "":
                        value = None
                elif isinstance(value, (int, float)):
                    # Keep as is for now
                    pass
                
                column_name = col_mapping.get(col, f"col_{col}")
                row_data[column_name] = value
            
            if has_data:
                data.append(row_data)
        
        return data
    
    def _infer_column_types(self, data: List[Dict[str, Any]]) -> Dict[str, str]:
        """Infer column types from data."""
        if not data:
            return {}
        
        column_types = {}
        
        for column in data[0].keys():
            values = [row.get(column) for row in data if row.get(column) is not None]
            
            if not values:
                column_types[column] = "string"
                continue
            
            # Check if all values are numeric
            numeric_count = sum(1 for v in values if isinstance(v, (int, float)))
            if numeric_count == len(values):
                column_types[column] = "float" if any(isinstance(v, float) for v in values) else "integer"
            else:
                column_types[column] = "string"
        
        return column_types
    
    def _generate_version(self, workbook_hash: str) -> str:
        """Generate version string from workbook hash and timestamp."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"{timestamp}_{workbook_hash[:8]}"
    
    def _create_extraction_report(self, metadata: Dict[str, Any], table_metadata: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """Create detailed extraction report."""
        report = {
            "extraction_summary": metadata,
            "tables_extracted": {},
            "defined_names_found": list(self.defined_names.keys()),
            "sheets_processed": self.workbook.sheetnames,
            "extraction_issues": []
        }
        
        for table_name, meta in table_metadata.items():
            report["tables_extracted"][table_name] = {
                "sheet": meta["sheet_name"],
                "range": meta["cell_range"],
                "rows": meta["row_count"],
                "columns": list(meta["column_types"].keys())
            }
        
        return report
